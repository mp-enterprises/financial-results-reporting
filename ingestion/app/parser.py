"""
Parser rozliczeń partnerskich (.xlsx).

Zasada nadrzędna: NIE parsujemy po numerach wierszy. Etykiety w arkuszach są
dopasowywane po znormalizowanym tekście, a tabele (Raport, Stok) po wykryciu
wiersza nagłówkowego. Dzięki temu wstawienie nowego wiersza w pliku przez
nadawcę nie psuje pipeline'u.

Wszystko, czego parser nie rozumie, kończy się wyjątkiem ParseError — plik
trafia do kwarantanny zamiast wgrać do bazy śmieci.
"""
from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

# --- oczekiwana struktura pliku --------------------------------------------
SHEET_HOWTO = "Jak czytać"
SHEET_KARTA = "Karta"
SHEET_KARTA_MJ = "Karta_MJ"
SHEET_RAPORT = "Raport"
SHEET_STOK = "Stok"
# Wymagane są tylko dwa arkusze — one niosą rozliczenie i rozbicie na SKU.
# Karta_MJ i Stok są opcjonalne: kanał MJ pojawia się nie w każdym okresie,
# a arkusz Stok nadawca dodał dopiero od 2026 M07 (starsze pliki go nie mają).
REQUIRED_SHEETS = {SHEET_KARTA, SHEET_RAPORT}
OPTIONAL_SHEETS = {SHEET_HOWTO, SHEET_KARTA_MJ, SHEET_STOK}
KNOWN_SHEETS = REQUIRED_SHEETS | OPTIONAL_SHEETS


# --- limity odporności na nieoczekiwaną zawartość --------------------------
# Nadawca może kiedyś dołożyć arkusz o dowolnej wielkości. Zrzut do raw jest
# cenny dla audytu, ale nie może wysadzić pamięci workera ani rozdąć bazy.
MAX_WIERSZY_ZRZUTU = 20_000      # na arkusz, w zrzucie do raw.sheet_payload
MAX_ARKUSZY = 40                 # powyżej tego plik jest podejrzany


class ParseError(ValueError):
    """Plik nie odpowiada oczekiwanej strukturze — nie wolno go załadować."""


# --- pomocnicze -------------------------------------------------------------
def normalize(text: Any) -> str:
    """Znormalizowana forma etykiety: bez ogonków, znaków interpunkcyjnych i
    wielokrotnych spacji. 'FV NI0 (prowizja = MAX(0; ...))' -> 'fv ni0 prowizja max 0'."""
    if text is None:
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("ł", "l").replace("Ł", "L")
    s = s.lower()
    s = re.sub(r"[^a-z0-9%]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def to_decimal(value: Any) -> Decimal | None:
    """Liczba z komórki -> Decimal. Obsługuje '—', '>999', '1 234,56', '12%'."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip()
    if s in {"", "-", "—", "–", "n/a", "N/A", "brak"}:
        return None
    capped = s.startswith(">")
    s = s.lstrip(">").lstrip("<").strip()
    s = s.replace("\xa0", "").replace(" ", "")
    pct = s.endswith("%")
    s = s.rstrip("%").replace(",", ".")
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    if pct:
        d = d / Decimal(100)
    return d


def to_int(value: Any) -> int | None:
    d = to_decimal(value)
    return int(d) if d is not None else None


def sha256_of(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# --- struktury danych -------------------------------------------------------
@dataclass
class SettlementFacts:
    """Odwzorowanie jednej karty rozliczeniowej (kanał MAIN albo MJ)."""
    channel_code: str
    koszty_ogolne: Decimal | None = None
    koszty_indywidualne: Decimal | None = None
    koszty_platformy: Decimal | None = None
    koszty_transport: Decimal | None = None
    zysk_ze_sprzedazy: Decimal | None = None
    zysk_po_kosztach: Decimal | None = None
    stawka_prowizji: Decimal | None = None
    prowizja_operatora: Decimal | None = None
    fv_bx: Decimal = Decimal(0)
    usluga_razem: Decimal | None = None
    zwroty_wartosciowe: Decimal = Decimal(0)
    fv_uslugowa: Decimal | None = None
    fv_partnera_netto: Decimal | None = None
    fv_partnera_szt: int | None = None
    kfv_netto: Decimal = Decimal(0)
    kfv_szt: int = 0
    saldo_towaru: Decimal | None = None
    srednia_cena_szt: Decimal | None = None
    do_zaplaty: Decimal | None = None
    prowizja_marketplace_pct: Decimal | None = None
    prowizja_techniczna_pct: Decimal | None = None
    przelicznik_ceny_fv: Decimal | None = None
    magazyn_wartosc: Decimal | None = None
    magazyn_szt: int | None = None
    magazyn_sku: int | None = None


@dataclass
class SalesLine:
    index_code: str
    product_name: str | None
    quantity: int
    unit_price: Decimal | None
    net_value: Decimal
    profit: Decimal
    line_no: int


@dataclass
class StockLine:
    index_code: str
    product_name: str | None
    qty_on_hand: int
    purchase_value: Decimal
    sales_month: int
    sales_3m: int
    sales_total: int
    avg_daily: Decimal | None
    days_cover: Decimal | None
    days_cover_capped: bool
    stock_status: str | None


@dataclass
class ParsedFile:
    partner_code: str
    period_year: int
    period_month: int
    generated_at: datetime | None
    settlements: dict[str, SettlementFacts]          # 'MAIN' / 'MJ'
    sales_lines: list[SalesLine]
    stock_lines: list[StockLine]
    raport_totals: dict[str, Any]
    notes: list[tuple[str, str | None, Decimal | None]]
    sheet_payloads: dict[str, list[list[Any]]] = field(default_factory=dict)
    unknown_sheets: list[str] = field(default_factory=list)
    # Nieblokujące spostrzeżenia, np. brak opcjonalnego arkusza. Trafiają do
    # wyniku ingestu, żeby operator wiedział, czego w pliku nie było.
    warnings: list[str] = field(default_factory=list)

    @property
    def has_stock(self) -> bool:
        return bool(self.stock_lines)


# --- parsowanie okresu i partnera -------------------------------------------
PERIOD_RE = re.compile(r"(20\d{2})\s*[-_ ]?\s*M\s*(\d{1,2})", re.IGNORECASE)
FILENAME_RE = re.compile(r"(20\d{2})M(\d{2})_([A-Za-z0-9]+)", re.IGNORECASE)


def parse_period(text: Any) -> tuple[int, int] | None:
    if text is None:
        return None
    m = PERIOD_RE.search(str(text))
    if not m:
        return None
    year, month = int(m.group(1)), int(m.group(2))
    if not 1 <= month <= 12:
        return None
    return year, month


def _rows(ws, limit: int | None = None) -> list[list[Any]]:
    """Wiersze arkusza; z limitem, gdy arkusz jest nieznany i może być ogromny."""
    out: list[list[Any]] = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if limit is not None and i >= limit:
            break
        out.append(list(r))
    return out


def _find_label(rows: Iterable[list[Any]], *candidates: str,
                value_col: int = 1, exact: bool = False) -> Any:
    """Zwraca wartość z kolumny value_col z pierwszego wiersza, którego
    etykieta (kolumna 0) pasuje do któregoś z kandydatów."""
    wanted = [normalize(c) for c in candidates]
    for row in rows:
        if not row:
            continue
        label = normalize(row[0])
        if not label:
            continue
        for w in wanted:
            if (label == w) if exact else (label.startswith(w) or w in label):
                if len(row) > value_col:
                    return row[value_col]
                return None
    return None


def _require(value: Any, label: str) -> Decimal:
    d = to_decimal(value)
    if d is None:
        raise ParseError(f"Brak wymaganej wartości liczbowej: {label!r}")
    return d


# --- karty rozliczeniowe ----------------------------------------------------
def parse_karta(rows: list[list[Any]]) -> SettlementFacts:
    f = SettlementFacts(channel_code="MAIN")
    g = lambda *lbl: _find_label(rows, *lbl)  # noqa: E731

    f.koszty_ogolne = to_decimal(g("koszty ogolne")) or Decimal(0)
    f.koszty_indywidualne = to_decimal(g("koszty indywidualne")) or Decimal(0)
    f.koszty_platformy = to_decimal(g("w tym platformy", "w tym prowizje platform"))
    f.koszty_transport = to_decimal(g("w tym transport"))
    f.zysk_ze_sprzedazy = to_decimal(g("zysk partnera ze sprzedazy"))
    f.zysk_po_kosztach = to_decimal(g("zysk partnera po kosztach"))
    f.stawka_prowizji = to_decimal(g("stawka prowizji nikcorp", "stawka prowizji"))
    f.prowizja_operatora = to_decimal(g("fv ni0"))
    f.fv_bx = to_decimal(g("fv z bx")) or Decimal(0)
    f.usluga_razem = to_decimal(g("usluga koszty ogolne"))
    f.zwroty_wartosciowe = to_decimal(g("zwroty wartosciowe")) or Decimal(0)
    f.fv_uslugowa = _require(g("fv uslugowa"), "FV usługowa")

    f.fv_partnera_netto = _require(g("fv partnera"), "FV partnera")
    f.kfv_netto = to_decimal(g("kfv partnera")) or Decimal(0)
    f.saldo_towaru = _require(g("saldo fv i kfv", "saldo fv kfv", "saldo"), "Saldo")
    f.srednia_cena_szt = to_decimal(g("srednia cena za szt"))
    f.do_zaplaty = _require(g("do zaplaty"), "Do zapłaty")

    f.prowizja_marketplace_pct = to_decimal(g("prowizja marketplace", "prowizja mp"))
    f.prowizja_techniczna_pct = to_decimal(g("prowizja techniczna"))
    f.przelicznik_ceny_fv = to_decimal(g("cena na fv cena sprzedazy"))

    f.magazyn_wartosc = to_decimal(g("wartosc magazynu"))
    f.magazyn_sku = to_int(g("pozycji sku na stanie"))

    # ilości sztuk siedzą w trzeciej kolumnie tych samych wierszy
    f.fv_partnera_szt = to_int(_find_label(rows, "fv partnera", value_col=2))
    f.kfv_szt = to_int(_find_label(rows, "kfv partnera", value_col=2)) or 0
    f.magazyn_szt = to_int(_find_label(rows, "wartosc magazynu", value_col=2))
    return f


def parse_karta_mj(rows: list[list[Any]]) -> SettlementFacts:
    f = SettlementFacts(channel_code="MJ")
    g = lambda *lbl: _find_label(rows, *lbl)  # noqa: E731

    f.koszty_ogolne = to_decimal(g("koszty ogolne")) or Decimal(0)
    f.koszty_indywidualne = to_decimal(g("koszty przydzielone", "razem koszty")) or Decimal(0)
    f.koszty_transport = to_decimal(g("koszty przydzielone transport", "koszty przydzielone"))
    f.zysk_ze_sprzedazy = to_decimal(g("zysk partnera ze sprzedazy"))
    f.zysk_po_kosztach = to_decimal(g("zysk po kosztach"))
    f.stawka_prowizji = to_decimal(g("stawka prowizji"))
    f.prowizja_operatora = to_decimal(g("prowizja 20", "prowizja"))
    f.usluga_razem = to_decimal(g("usluga mj"))
    f.fv_uslugowa = f.usluga_razem            # w kanale MJ usługa == cała faktura

    f.fv_partnera_netto = _require(g("fv partnera"), "FV partnera (MJ)")
    f.kfv_netto = to_decimal(g("kfv partnera")) or Decimal(0)
    f.saldo_towaru = _require(g("saldo"), "Saldo (MJ)")
    f.srednia_cena_szt = to_decimal(g("srednia cena za szt"))
    f.do_zaplaty = _require(g("do zaplaty"), "Do zapłaty (MJ)")

    f.prowizja_marketplace_pct = to_decimal(g("prowizja mp", "prowizja marketplace"))
    f.prowizja_techniczna_pct = to_decimal(g("prowizja techniczna"))
    f.przelicznik_ceny_fv = to_decimal(g("cena na fv cena sprzedazy"))

    f.fv_partnera_szt = to_int(_find_label(rows, "fv partnera", value_col=2))
    f.kfv_szt = to_int(_find_label(rows, "kfv partnera", value_col=2)) or 0
    return f


# --- tabele -----------------------------------------------------------------
# Mapy kolumn: klucz logiczny -> lista dopuszczalnych etykiet (znormalizowanych).
# Kolejność w liście to kolejność dopasowania; pierwsza pasująca wygrywa.
KOLUMNY_RAPORT: dict[str, list[str]] = {
    "index_code":   ["indeks", "indeks towaru", "sku", "kod"],
    "product_name": ["nazwa", "nazwa towaru", "opis"],
    "quantity":     ["ilosc", "ilosc szt", "sztuk"],
    "unit_price":   ["cena", "cena srednia", "srednia cena"],
    "net_value":    ["wartosc", "wartosc netto", "obrot"],
    "profit":       ["zysk", "zysk netto", "marza"],
}
WYMAGANE_RAPORT = ["index_code", "net_value", "profit"]

KOLUMNY_STOK: dict[str, list[str]] = {
    "index_code":     ["indeks", "indeks towaru", "sku", "kod"],
    "product_name":   ["nazwa", "nazwa towaru", "opis"],
    "qty_on_hand":    ["stan szt", "stan", "ilosc", "stan magazynowy"],
    "purchase_value": ["wartosc zakupu", "wartosc", "wartosc magazynu"],
    "sales_month":    ["sprzedaz w miesiacu", "sprzedaz miesiac", "sprzedaz 1 mies"],
    "sales_3m":       ["sprzedaz 3 mies", "sprzedaz 3m", "sprzedaz kwartal"],
    "sales_total":    ["sprzedaz lacznie", "sprzedaz razem", "sprzedaz calkowita"],
    "avg_daily":      ["sr dziennie", "srednia dziennie", "srednio dziennie"],
    "days_cover":     ["wystarczy na dni", "wystarczy na", "dni pokrycia"],
    "stock_status":   ["status", "status rotacji"],
}
WYMAGANE_STOK = ["index_code", "qty_on_hand", "purchase_value"]


def _mapuj_kolumny(naglowek: list[Any], definicje: dict[str, list[str]],
                   wymagane: list[str], arkusz: str,
                   ostrzezenia: list[str]) -> dict[str, int]:
    """Buduje mapę: klucz logiczny -> indeks kolumny, na podstawie ETYKIET.

    To jest zabezpieczenie przed najgroźniejszą zmianą, jaką nadawca może
    wprowadzić: wstawieniem albo przestawieniem kolumny. Odczyt po stałych
    pozycjach przesunąłby wtedy wszystkie wartości i załadował do bazy ciche
    przekłamania — cena trafiłaby do wartości, wartość do zysku.
    """
    etykiety = {}
    for idx, komorka in enumerate(naglowek):
        klucz = normalize(komorka)
        if klucz and klucz not in etykiety:
            etykiety[klucz] = idx

    mapa: dict[str, int] = {}
    for logiczna, warianty in definicje.items():
        for wariant in warianty:                       # najpierw trafienie dokładne
            if wariant in etykiety:
                mapa[logiczna] = etykiety[wariant]
                break
        else:
            for wariant in warianty:                   # potem po przedrostku
                trafienia = [i for e, i in etykiety.items() if e.startswith(wariant)]
                if len(trafienia) == 1:
                    mapa[logiczna] = trafienia[0]
                    break

    brakujace = [k for k in wymagane if k not in mapa]
    if brakujace:
        raise ParseError(
            f"Arkusz {arkusz}: nie rozpoznano wymaganych kolumn {brakujace}. "
            f"Znalezione nagłówki: {sorted(etykiety)}"
        )

    opcjonalne_brak = [k for k in definicje if k not in mapa]
    if opcjonalne_brak:
        ostrzezenia.append(
            f"Arkusz {arkusz}: brak opcjonalnych kolumn {opcjonalne_brak} — "
            f"te pola zostaną puste."
        )

    nadmiarowe = sorted(set(etykiety) - {w for v in definicje.values() for w in v})
    if nadmiarowe:
        ostrzezenia.append(
            f"Arkusz {arkusz}: nowe, nieużywane kolumny {nadmiarowe} — pominięte."
        )
    return mapa


def _kom(row: list[Any], mapa: dict[str, int], klucz: str) -> Any:
    """Wartość komórki po kluczu logicznym; None gdy kolumny nie ma w pliku."""
    idx = mapa.get(klucz)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _find_header_row(rows: list[list[Any]], first_cell: str, min_cols: int) -> int:
    target = normalize(first_cell)
    for i, row in enumerate(rows):
        if row and normalize(row[0]) == target and sum(c is not None for c in row) >= min_cols:
            return i
    raise ParseError(f"Nie znaleziono wiersza nagłówkowego zaczynającego się od {first_cell!r}")


def parse_raport(rows: list[list[Any]],
                 ostrzezenia: list[str] | None = None) -> tuple[list[SalesLine], dict[str, Any]]:
    ostrzezenia = ostrzezenia if ostrzezenia is not None else []
    header_idx = _find_header_row(rows, "Indeks", 5)
    mapa = _mapuj_kolumny(rows[header_idx], KOLUMNY_RAPORT, WYMAGANE_RAPORT,
                          "Raport", ostrzezenia)

    # Suma kontrolna „RAZEM” stoi nad nagłówkiem, więc trafia w te same kolumny.
    totals: dict[str, Any] = {}
    for row in rows[:header_idx]:
        if row and normalize(row[0]) == "razem":
            totals = {
                "quantity":  to_int(_kom(row, mapa, "quantity")),
                "net_value": to_decimal(_kom(row, mapa, "net_value")),
                "profit":    to_decimal(_kom(row, mapa, "profit")),
            }
            break

    lines: list[SalesLine] = []
    for n, row in enumerate(rows[header_idx + 1:], start=1):
        if not row:
            continue
        surowy_kod = _kom(row, mapa, "index_code")
        if surowy_kod is None or str(surowy_kod).strip() == "":
            continue
        code = str(surowy_kod).strip()
        if normalize(code) in {"razem", "suma", "total"}:
            continue

        net = to_decimal(_kom(row, mapa, "net_value"))
        profit = to_decimal(_kom(row, mapa, "profit"))
        if net is None and profit is None:
            continue

        nazwa = _kom(row, mapa, "product_name")
        lines.append(SalesLine(
            index_code=code,
            product_name=str(nazwa).strip() if nazwa is not None else None,
            quantity=to_int(_kom(row, mapa, "quantity")) or 0,
            unit_price=to_decimal(_kom(row, mapa, "unit_price")),
            net_value=net if net is not None else Decimal(0),
            profit=profit if profit is not None else Decimal(0),
            line_no=n,
        ))
    if not lines:
        raise ParseError("Arkusz Raport nie zawiera żadnych pozycji sprzedaży")
    return lines, totals


def parse_stok(rows: list[list[Any]],
               ostrzezenia: list[str] | None = None) -> list[StockLine]:
    ostrzezenia = ostrzezenia if ostrzezenia is not None else []
    header_idx = _find_header_row(rows, "Indeks", 8)
    mapa = _mapuj_kolumny(rows[header_idx], KOLUMNY_STOK, WYMAGANE_STOK,
                          "Stok", ostrzezenia)

    lines: list[StockLine] = []
    for row in rows[header_idx + 1:]:
        if not row:
            continue
        surowy_kod = _kom(row, mapa, "index_code")
        if surowy_kod is None or str(surowy_kod).strip() == "":
            continue
        code = str(surowy_kod).strip()
        if normalize(code) in {"razem", "suma", "total"}:
            continue

        raw_cover = _kom(row, mapa, "days_cover")
        capped = isinstance(raw_cover, str) and raw_cover.strip().startswith(">")
        nazwa = _kom(row, mapa, "product_name")
        status = _kom(row, mapa, "stock_status")
        lines.append(StockLine(
            index_code=code,
            product_name=str(nazwa).strip() if nazwa is not None else None,
            qty_on_hand=to_int(_kom(row, mapa, "qty_on_hand")) or 0,
            purchase_value=to_decimal(_kom(row, mapa, "purchase_value")) or Decimal(0),
            sales_month=to_int(_kom(row, mapa, "sales_month")) or 0,
            sales_3m=to_int(_kom(row, mapa, "sales_3m")) or 0,
            sales_total=to_int(_kom(row, mapa, "sales_total")) or 0,
            avg_daily=to_decimal(_kom(row, mapa, "avg_daily")),
            days_cover=to_decimal(raw_cover),
            days_cover_capped=capped,
            stock_status=str(status).strip() if status is not None else None,
        ))
    if not lines:
        raise ParseError("Arkusz Stok nie zawiera żadnych pozycji")
    return lines


def parse_notes(rows: list[list[Any]]) -> tuple[list[tuple[str, str | None, Decimal | None]], datetime | None]:
    notes: list[tuple[str, str | None, Decimal | None]] = []
    generated_at: datetime | None = None
    seen: set[str] = set()
    for row in rows:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        key = normalize(label)
        if not key or key in seen:
            continue
        value = row[1] if len(row) > 1 else None
        num = to_decimal(value)
        if key.startswith("wygenerowano"):
            raw = value if value is not None else label.split(":", 1)[-1]
            if isinstance(raw, datetime):
                generated_at = raw
            else:
                try:
                    generated_at = datetime.strptime(str(raw).strip(), "%Y-%m-%d %H:%M")
                except ValueError:
                    generated_at = None
        if value is None and num is None:
            continue
        seen.add(key)
        notes.append((key[:200], None if value is None else str(value)[:500], num))
    return notes, generated_at


# --- wejście publiczne ------------------------------------------------------
def parse_workbook(path: str | Path, file_name: str | None = None) -> ParsedFile:
    path = Path(path)
    file_name = file_name or path.name
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"Nie można otworzyć pliku jako XLSX: {exc}") from exc

    try:
        present = set(wb.sheetnames)
        missing = REQUIRED_SHEETS - present
        if missing:
            raise ParseError(
                f"Brak wymaganych arkuszy: {sorted(missing)}. Znalezione: {sorted(present)}"
            )

        nieznane = [n for n in wb.sheetnames if n not in KNOWN_SHEETS]

        # Arkusze, które znamy, czytamy w całości — na nich stoi rozliczenie.
        # Nieznane czytamy z limitem: trafiają wyłącznie do zrzutu audytowego,
        # więc nie ma powodu wciągać do pamięci arkusza o dowolnej wielkości.
        sheets = {}
        for name in wb.sheetnames:
            limit = None if name in KNOWN_SHEETS else MAX_WIERSZY_ZRZUTU
            sheets[name] = _rows(wb[name], limit)
    finally:
        wb.close()

    karta_rows = sheets[SHEET_KARTA]

    # --- partner i okres: karta -> "Jak czytać" -> nazwa pliku ---
    partner = _find_label(karta_rows, "partner")
    period_raw = _find_label(karta_rows, "okres")
    if SHEET_HOWTO in sheets:
        partner = partner or _find_label(sheets[SHEET_HOWTO], "partner")
        period_raw = period_raw or _find_label(sheets[SHEET_HOWTO], "okres")
    period = parse_period(period_raw)

    if partner is None or period is None:
        m = FILENAME_RE.search(file_name)
        if not m:
            raise ParseError(
                f"Nie udało się ustalić partnera/okresu ani z zawartości, ani z nazwy pliku {file_name!r}"
            )
        period = period or (int(m.group(1)), int(m.group(2)))
        partner = partner or m.group(3)

    partner_code = str(partner).strip().upper()
    year, month = period

    # --- karty ---
    settlements = {"MAIN": parse_karta(karta_rows)}
    if SHEET_KARTA_MJ in sheets:
        mj_rows = sheets[SHEET_KARTA_MJ]
        if to_decimal(_find_label(mj_rows, "fv partnera")) is not None:
            settlements["MJ"] = parse_karta_mj(mj_rows)

    warnings: list[str] = []

    # Nowy arkusz od nadawcy NIE jest błędem — zostaje odnotowany i zarchiwizowany
    # w raw.sheet_payload, żeby dało się go później przeanalizować i ewentualnie
    # dopisać obsługę. Rozliczenie liczy się z arkuszy, które znamy.
    if nieznane:
        warnings.append(
            f"Plik zawiera nowe arkusze: {nieznane}. Nie wpływają na rozliczenie, "
            f"ale zostały zarchiwizowane — sprawdź, czy nadawca nie zmienił formatu."
        )
    if len(sheets) > MAX_ARKUSZY:
        warnings.append(
            f"Plik ma {len(sheets)} arkuszy (spodziewane do {MAX_ARKUSZY}) — "
            f"to nietypowe, warto zweryfikować źródło."
        )
    for nazwa, wiersze in sheets.items():
        if nazwa not in KNOWN_SHEETS and len(wiersze) >= MAX_WIERSZY_ZRZUTU:
            warnings.append(
                f"Arkusz '{nazwa}' przekroczył {MAX_WIERSZY_ZRZUTU} wierszy — "
                f"w archiwum zapisano tylko początek."
            )

    sales_lines, raport_totals = parse_raport(sheets[SHEET_RAPORT], warnings)

    if SHEET_STOK in sheets:
        stock_lines = parse_stok(sheets[SHEET_STOK], warnings)
    else:
        stock_lines = []
        warnings.append(
            "Plik nie zawiera arkusza 'Stok' — dane magazynowe za ten okres nie "
            "zostaną wczytane. Starsze rozliczenia (przed 2026 M07) go nie mają."
        )
    if "MJ" not in settlements:
        warnings.append("Plik nie zawiera rozliczenia kanału MJ / Amazon.")

    notes, generated_at = parse_notes(sheets.get(SHEET_HOWTO, []))

    return ParsedFile(
        partner_code=partner_code,
        period_year=year,
        period_month=month,
        generated_at=generated_at,
        settlements=settlements,
        sales_lines=sales_lines,
        stock_lines=stock_lines,
        raport_totals=raport_totals,
        notes=notes,
        sheet_payloads={
            name: [[_jsonable(c) for c in row] for row in rows[:MAX_WIERSZY_ZRZUTU]]
            for name, rows in sheets.items()
        },
        unknown_sheets=sorted(set(sheets) - KNOWN_SHEETS),
        warnings=warnings,
    )


def _jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value
