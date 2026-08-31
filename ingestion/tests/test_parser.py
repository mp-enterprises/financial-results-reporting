"""
Testy parsera — najbardziej kruchego elementu systemu.

Uruchomienie:  pytest ingestion/tests -q
"""
from __future__ import annotations

import shutil
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.parser import (MAX_WIERSZY_ZRZUTU, ParseError, normalize, parse_period,
                        parse_workbook,
                        sha256_of, to_decimal)

SAMPLE = Path(__file__).resolve().parents[2] / "sample" / "Rozliczenie_2026M07_ZZMP1.xlsx"
pytestmark = pytest.mark.skipif(not SAMPLE.exists(), reason="brak pliku przykładowego")


# --- funkcje pomocnicze -----------------------------------------------------
@pytest.mark.parametrize("raw,expected", [
    ("1234,56", Decimal("1234.56")),
    ("1 234.56", Decimal("1234.56")),
    ("\xa01234.56", Decimal("1234.56")),
    ("12%", Decimal("0.12")),
    (">999", Decimal("999")),
    ("—", None), ("", None), (None, None), ("brak", None),
    (1234.5, Decimal("1234.5")),
])
def test_to_decimal(raw, expected):
    assert to_decimal(raw) == expected


def test_normalize_usuwa_polskie_znaki():
    assert normalize("Zysk partnera (po kosztach)") == "zysk partnera po kosztach"
    assert normalize("KOSZTY OGÓLNE") == "koszty ogolne"
    assert normalize("Wartość magazynu — RAZEM") == "wartosc magazynu razem"


@pytest.mark.parametrize("raw,expected", [
    ("2026  M07", (2026, 7)), ("2026 M7", (2026, 7)),
    ("2025M12", (2025, 12)), ("bez okresu", None), ("2026 M13", None),
])
def test_parse_period(raw, expected):
    assert parse_period(raw) == expected


# --- parsowanie pliku wzorcowego -------------------------------------------
@pytest.fixture(scope="module")
def parsed():
    return parse_workbook(SAMPLE)


def test_naglowek(parsed):
    assert parsed.partner_code == "ZZMP1"
    assert (parsed.period_year, parsed.period_month) == (2026, 7)
    assert parsed.generated_at is not None


def test_kanaly(parsed):
    assert set(parsed.settlements) == {"MAIN", "MJ"}


def test_kwoty_kanalu_glownego(parsed):
    s = parsed.settlements["MAIN"]
    assert s.do_zaplaty == Decimal("60041.44380764045")
    assert s.saldo_towaru == Decimal("126459.39279999997")
    assert s.fv_partnera_szt == 1261
    assert s.kfv_szt == 104
    assert s.stawka_prowizji == Decimal("0.2")
    assert s.magazyn_sku == 1096


def test_arytmetyka_sie_spina(parsed):
    """Najważniejszy test biznesowy: rachunek z pliku musi się zgadzać."""
    for channel, s in parsed.settlements.items():
        assert abs(s.fv_partnera_netto - s.kfv_netto - s.saldo_towaru) < Decimal("0.01"), channel
        assert abs(s.saldo_towaru - s.fv_uslugowa - s.do_zaplaty) < Decimal("0.01"), channel
        expected_fee = max(Decimal(0), s.zysk_po_kosztach * s.stawka_prowizji)
        assert abs(expected_fee - s.prowizja_operatora) < Decimal("0.01"), channel


def test_tabele(parsed):
    assert len(parsed.sales_lines) == 306
    assert len(parsed.stock_lines) == 1096
    assert len({l.index_code for l in parsed.sales_lines}) == 306
    assert len({l.index_code for l in parsed.stock_lines}) == 1096


def test_sumy_z_raportu(parsed):
    suma = sum(l.net_value for l in parsed.sales_lines)
    assert abs(suma - parsed.raport_totals["net_value"]) < Decimal("0.01")
    zysk = sum(l.profit for l in parsed.sales_lines)
    assert abs(zysk - parsed.raport_totals["profit"]) < Decimal("0.01")


def test_stok_zgodny_z_karta(parsed):
    suma = sum(l.purchase_value for l in parsed.stock_lines)
    assert abs(suma - parsed.settlements["MAIN"].magazyn_wartosc) < Decimal("0.01")
    assert sum(l.qty_on_hand for l in parsed.stock_lines) == parsed.settlements["MAIN"].magazyn_szt


def test_pozycje_ujemne_sa_zachowane(parsed):
    """Zwroty przewyższające sprzedaż dają ujemne ilości — nie wolno ich gubić."""
    assert any(l.quantity < 0 for l in parsed.sales_lines)
    assert any(l.profit < 0 for l in parsed.sales_lines)


def test_martwy_stok_ma_pusty_wskaznik(parsed):
    dead = [l for l in parsed.stock_lines if l.stock_status == "martwy stok"]
    assert dead and all(l.days_cover is None for l in dead)


def test_capped_days_cover(parsed):
    """'>999' w pliku musi być oznaczone flagą, a nie zamienione w NULL."""
    capped = [l for l in parsed.stock_lines if l.days_cover_capped]
    assert capped and all(l.days_cover == Decimal("999") for l in capped)


# --- odporność --------------------------------------------------------------
def test_przesuniecie_wierszy_nie_psuje_parsowania(tmp_path):
    """Nadawca wstawia dodatkowy wiersz na górze Karty — parser ma to przetrwać."""
    dst = tmp_path / "przesuniety.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    wb["Karta"].insert_rows(1, 3)
    wb["Karta"]["A1"] = "NOWY NAGŁÓWEK DODANY PRZEZ NADAWCĘ"
    wb["Raport"].insert_rows(1, 2)
    wb.save(dst)

    p = parse_workbook(dst)
    assert p.settlements["MAIN"].do_zaplaty == Decimal("60041.44380764045")
    assert len(p.sales_lines) == 306


def test_brak_arkusza_stok_jest_dozwolony(tmp_path):
    """Arkusz Stok nadawca dodał dopiero od 2026 M07 — starsze pliki go nie mają
    i muszą się wczytać, tylko bez danych magazynowych."""
    dst = tmp_path / "bez_stoku.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    del wb["Stok"]
    wb.save(dst)

    p = parse_workbook(dst)
    assert p.stock_lines == []
    assert p.has_stock is False
    assert any("Stok" in w for w in p.warnings)
    # rozliczenie musi się wczytać bez zmian
    assert p.settlements["MAIN"].do_zaplaty == Decimal("60041.44380764045")
    assert len(p.sales_lines) == 306


def test_brak_arkusza_karta_konczy_sie_bledem(tmp_path):
    """Karta i Raport są wymagane — bez nich nie ma czego ładować."""
    dst = tmp_path / "bez_karty.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    del wb["Karta"]
    wb.save(dst)
    with pytest.raises(ParseError, match="Brak wymaganych arkuszy"):
        parse_workbook(dst)


# --- plik z poprzedniego miesiąca (inna struktura: bez arkusza Stok) --------
SAMPLE_06 = SAMPLE.parent / "Rozliczenie_2026M06_ZZMP1.xlsx"


@pytest.mark.skipif(not SAMPLE_06.exists(), reason="brak pliku za 2026-06")
def test_plik_czerwcowy_bez_stoku():
    p = parse_workbook(SAMPLE_06)
    assert p.partner_code == "ZZMP1"
    assert (p.period_year, p.period_month) == (2026, 6)
    assert set(p.settlements) == {"MAIN", "MJ"}
    assert len(p.sales_lines) == 235
    assert p.stock_lines == []
    assert p.settlements["MAIN"].do_zaplaty == Decimal("63869.87921239117")
    assert p.settlements["MJ"].do_zaplaty == Decimal("1479.2245031855782")
    # w Karcie bez sekcji magazynowej te pola muszą zostać puste, nie wyzerowane
    assert p.settlements["MAIN"].magazyn_wartosc is None
    assert p.settlements["MAIN"].magazyn_sku is None


@pytest.mark.skipif(not SAMPLE_06.exists(), reason="brak pliku za 2026-06")
def test_arytmetyka_czerwca_sie_spina():
    p = parse_workbook(SAMPLE_06)
    for channel, s in p.settlements.items():
        assert abs(s.fv_partnera_netto - s.kfv_netto - s.saldo_towaru) < Decimal("0.01"), channel
        assert abs(s.saldo_towaru - s.fv_uslugowa - s.do_zaplaty) < Decimal("0.01"), channel
    suma = sum(l.net_value for l in p.sales_lines)
    assert abs(suma - p.raport_totals["net_value"]) < Decimal("0.01")


def test_pusty_plik_konczy_sie_bledem(tmp_path):
    dst = tmp_path / "pusty.xlsx"
    wb = Workbook()
    wb.active.title = "Arkusz1"
    wb.save(dst)
    with pytest.raises(ParseError):
        parse_workbook(dst)


def test_uszkodzony_plik_konczy_sie_bledem(tmp_path):
    dst = tmp_path / "uszkodzony.xlsx"
    dst.write_bytes(b"to nie jest xlsx")
    with pytest.raises(ParseError, match="Nie można otworzyć"):
        parse_workbook(dst)


def test_partner_z_nazwy_pliku_gdy_brak_w_tresci(tmp_path):
    dst = tmp_path / "Rozliczenie_2025M03_TESTPARTNER.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    for sheet in ("Karta", "Jak czytać"):
        for row in wb[sheet].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower().startswith(("partner", "okres")):
                    cell.value = "usunięte"
    wb.save(dst)
    p = parse_workbook(dst)
    assert p.partner_code == "TESTPARTNER"
    assert (p.period_year, p.period_month) == (2025, 3)


def test_sha256_jest_stabilny():
    assert sha256_of(SAMPLE) == sha256_of(SAMPLE)
    assert len(sha256_of(SAMPLE)) == 64


# =============================================================================
# Odporność na zmiany formatu po stronie nadawcy
#
# Najgroźniejsza zmiana to nie nowy arkusz, tylko wstawiona lub przestawiona
# KOLUMNA: odczyt po stałych pozycjach przesunąłby wartości i załadował do bazy
# ciche przekłamania (cena jako wartość, wartość jako zysk). Dlatego kolumny są
# mapowane po etykietach nagłówka, a te testy tego pilnują.
# =============================================================================

WZORZEC_PIERWSZEJ_POZYCJI = dict(
    index_code="DB000003", quantity=71,
    net_value=Decimal("16937.86"), profit=Decimal("12974.64"),
)


def _sprawdz_pierwsza_pozycje(parsed):
    linia = parsed.sales_lines[0]
    for pole, oczekiwane in WZORZEC_PIERWSZEJ_POZYCJI.items():
        assert getattr(linia, pole) == oczekiwane, pole


def test_nowy_arkusz_nie_psuje_parsera(tmp_path):
    """Nadawca dokłada arkusz — rozliczenie ma się wczytać bez zmian."""
    dst = tmp_path / "nowy_arkusz.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    ws = wb.create_sheet("Prowizje_2027")
    ws["A1"] = "Nowa tabela nadawcy"
    for i in range(2, 60):
        ws.cell(row=i, column=1, value=f"poz{i}")
        ws.cell(row=i, column=2, value=i * 1.5)
    wb.save(dst)

    p = parse_workbook(dst)
    _sprawdz_pierwsza_pozycje(p)
    assert p.settlements["MAIN"].do_zaplaty == Decimal("60041.44380764045")
    assert "Prowizje_2027" in p.unknown_sheets
    assert any("nowe arkusze" in w for w in p.warnings)
    assert "Prowizje_2027" in p.sheet_payloads        # zarchiwizowany do audytu


def test_wiele_nowych_arkuszy(tmp_path):
    dst = tmp_path / "wiele.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    for nazwa in ("Aneks", "Korekty_VAT", "Notatki"):
        wb.create_sheet(nazwa)["A1"] = nazwa
    wb.save(dst)

    p = parse_workbook(dst)
    _sprawdz_pierwsza_pozycje(p)
    assert len(p.unknown_sheets) == 3


def test_ogromny_nowy_arkusz_jest_przycinany(tmp_path):
    """Nowy arkusz o dowolnej wielkości nie może wysadzić pamięci workera."""
    dst = tmp_path / "ogromny.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    ws = wb.create_sheet("Log")
    for i in range(1, MAX_WIERSZY_ZRZUTU + 5_001):
        ws.cell(row=i, column=1, value=i)
    wb.save(dst)

    p = parse_workbook(dst)
    _sprawdz_pierwsza_pozycje(p)
    assert len(p.sheet_payloads["Log"]) <= MAX_WIERSZY_ZRZUTU
    assert any("przekroczy" in w for w in p.warnings)


def test_wstawiona_kolumna_w_raporcie(tmp_path):
    """Kolumna wstawiona w środku NIE MOŻE przesunąć wartości."""
    dst = tmp_path / "kolumna.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    ws = wb["Raport"]
    ws.insert_cols(2)
    ws.cell(row=3, column=2).value = "EAN"
    for i in range(4, ws.max_row + 1):
        ws.cell(row=i, column=2, value="590123456789")
    wb.save(dst)

    p = parse_workbook(dst)
    _sprawdz_pierwsza_pozycje(p)
    assert any("ean" in w.lower() for w in p.warnings)


def test_przestawione_kolumny_w_stoku(tmp_path):
    dst = tmp_path / "stok_kol.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    ws = wb["Stok"]
    naglowek = [ws.cell(row=3, column=c).value for c in range(1, 11)]
    dane = [[ws.cell(row=r, column=c).value for c in range(1, 11)]
            for r in range(4, ws.max_row + 1)]
    kolejnosc = [0, 3, 1, 2, 9, 4, 5, 6, 7, 8]
    for c, src in enumerate(kolejnosc, start=1):
        ws.cell(row=3, column=c, value=naglowek[src])
        for r, wiersz in enumerate(dane, start=4):
            ws.cell(row=r, column=c, value=wiersz[src])
    wb.save(dst)

    p = parse_workbook(dst)
    poz = next(l for l in p.stock_lines if l.index_code == "YXT003")
    assert poz.qty_on_hand == 78
    assert poz.purchase_value == Decimal("23772.84")
    assert poz.stock_status == "OK"
    suma = sum(l.purchase_value for l in p.stock_lines)
    assert abs(suma - Decimal("269484.89")) < Decimal("0.01")


def test_brak_wymaganej_kolumny_konczy_sie_bledem(tmp_path):
    """Lepiej kwarantanna niż ciche śmieci w rozliczeniu."""
    dst = tmp_path / "brak_kol.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    wb["Raport"].delete_cols(5)                       # kolumna Wartość
    wb.save(dst)
    with pytest.raises(ParseError, match="nie rozpoznano wymaganych kolumn"):
        parse_workbook(dst)


def test_brak_opcjonalnej_kolumny_daje_ostrzezenie(tmp_path):
    dst = tmp_path / "brak_opc.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    wb["Stok"].delete_cols(10)                        # kolumna Status
    wb.save(dst)

    p = parse_workbook(dst)
    assert len(p.stock_lines) == 1096
    assert all(l.stock_status is None for l in p.stock_lines)
    assert any("opcjonalnych kolumn" in w for w in p.warnings)


def test_kombinacja_zmian(tmp_path):
    """Nowy arkusz + nowa kolumna + nowe wiersze naraz."""
    dst = tmp_path / "kombo.xlsx"
    shutil.copy(SAMPLE, dst)
    wb = load_workbook(dst)
    wb.create_sheet("Zalacznik")["A1"] = "x"
    wb["Karta"].insert_rows(1, 2)
    wb["Karta"]["A1"] = "NAGŁÓWEK DODANY PRZEZ NADAWCĘ"
    ws = wb["Raport"]
    ws.insert_cols(3)
    ws.cell(row=3, column=3).value = "Kategoria"
    wb.save(dst)

    p = parse_workbook(dst)
    _sprawdz_pierwsza_pozycje(p)
    assert p.settlements["MAIN"].do_zaplaty == Decimal("60041.44380764045")
    assert len(p.sales_lines) == 306
